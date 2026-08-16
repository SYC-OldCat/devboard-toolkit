"""Excel 结果报告生成

处理完成后生成两个 Excel:
- 操作成功.xlsx: 成功复制的问题列表
- 操作失败.xlsx: 失败的问题列表 (路径未找到 / 复制异常等)

格式:
- 表头: 问题ID / 标题 / 数据路径
- 表头蓝底白字居中, 数据左对齐, 全表加边框
- 列宽统一 40
"""

import os
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = ["问题ID", "标题", "数据路径", "失败原因"]


def _build_sheet(ws, results: List[Dict[str, Any]]):
    """填充一个 sheet (表头 + 数据)"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头
    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 数据
    data_alignment = Alignment(vertical="center", wrap_text=True)
    for row, item in enumerate(results, 2):
        issue_id = item.get("issue_id", "")
        title = item.get("summary", "")
        video_path = item.get("video_path", "")
        fail_reason = item.get("fail_reason", "")

        for col, val in enumerate([issue_id, title, video_path, fail_reason], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = data_alignment

    # 列宽
    widths = [18, 40, 60, 30]
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]


def generate_excel_results(results: List[Dict[str, Any]], output_dir: str):
    """根据处理结果生成成功/失败两个 Excel

    Args:
        results: 每条记录含 issue_id / summary / video_path / success
        output_dir: Excel 输出目录
    """
    success_results = [r for r in results if r.get("success", False)]
    failed_results = [r for r in results if not r.get("success", False)]

    import datetime as _dt

    def _safe_save(wb_: openpyxl.Workbook, output_dir_: str, base_name: str) -> str:
        """保存 Excel,如果目标被占用 (Permission denied) 则改用带时间戳的备用名"""
        target = os.path.join(output_dir_, base_name + ".xlsx")
        try:
            wb_.save(target)
            return target
        except PermissionError:
            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            alt = os.path.join(output_dir_, f"{base_name}_{ts}.xlsx")
            wb_.save(alt)
            print(f"  [!] {os.path.basename(target)} 被占用 (可能 Excel/WPS 打开中), 已改为: {alt}")
            return alt

    if success_results:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "操作成功"
        _build_sheet(ws, success_results)
        path = _safe_save(wb, output_dir, "操作成功")
        print(f"\n[+] 已生成成功结果文件: {path}")

    if failed_results:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "操作失败"
        _build_sheet(ws, failed_results)
        path = _safe_save(wb, output_dir, "操作失败")
        print(f"\n[+] 已生成失败结果文件: {path}")
